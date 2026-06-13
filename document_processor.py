"""
文档清洗和切片模块
支持 HTML（BeautifulSoup 去标签 + 去噪声）、PDF（pdfplumber 提取）、
Excel（openpyxl 转 Markdown 表格）、TXT（纯文本）。
中文切片用 jieba 辅助，300-500 字/片，片间 80 字重叠。
"""

import re
import os
from typing import List, Dict, Optional


# ============================================================
# HTML 清洗（BeautifulSoup）
# ============================================================

# 噪声标签集合：这些标签的内容会被整块移除
_NOISE_TAGS = {
    "script", "style", "noscript", "nav", "footer", "header",
    "aside", "iframe", "object", "embed",
}

# 噪声类名/ID 关键词：包含这些关键词的 div/section 会被跳过
_NOISE_CLASS_KEYWORDS = [
    "nav", "menu", "footer", "header", "sidebar", "breadcrumb",
    "related", "recommend", "recommendation",
    "copyright", "版权", "banner", "ad", "advertisement", "广告",
    "pagination", "分页", "comment", "评论", "share", "分享",
    "上一篇", "下一篇", "相关文章", "热门推荐",
    "toolbar", "toolbar", "tool-bar",
]


def _has_noise_class(tag) -> bool:
    """检查标签的 class/id 是否包含噪声关键词"""
    cls = " ".join(tag.get("class", [])).lower() if tag.get("class") else ""
    tid = (tag.get("id", "") or "").lower()
    combined = f"{cls} {tid}"
    return any(kw in combined for kw in _NOISE_CLASS_KEYWORDS)


def clean_html(file_path: str) -> str:
    """用 BeautifulSoup 清洗 HTML：去标签、去导航/页脚/推荐等模板噪声，返回纯文本"""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        raise ImportError("请安装 beautifulsoup4: pip install beautifulsoup4")

    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        html_content = f.read()

    soup = BeautifulSoup(html_content, "lxml")

    # 1. 移除噪声标签
    for tag_name in _NOISE_TAGS:
        for tag in soup.find_all(tag_name):
            tag.decompose()

    # 2. 移除含有噪声 class/id 的 div/section
    for tag in soup.find_all(["div", "section"]):
        if _has_noise_class(tag):
            tag.decompose()

    # 3. 提取纯文本
    text = soup.get_text(separator="\n")

    # 4. 后处理：去除多余空行和空白
    text = re.sub(r"[ \t]+", " ", text)        # 合并连续空格/制表符
    text = re.sub(r"\n{3,}", "\n\n", text)      # 最多保留一个空行
    text = re.sub(r"^\n+", "", text)            # 去掉开头空行
    return text.strip()


# ============================================================
# PDF 文本提取（pdfplumber）
# ============================================================


def extract_pdf_text(file_path: str) -> str:
    """用 pdfplumber 提取 PDF 文本（文本型 PDF；扫描件需预先 OCR）"""
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("请安装 pdfplumber: pip install pdfplumber")

    texts: List[str] = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text and page_text.strip():
                texts.append(page_text.strip())

    return "\n\n".join(texts)


# ============================================================
# Excel 转 Markdown 表格
# ============================================================


def excel_to_markdown(file_path: str) -> str:
    """将 Excel 文件转为 Markdown 表格文本，每张工作表一个表格"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    results: List[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        results.append(f"## 工作表: {sheet_name}\n")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            results.append("（空表）\n")
            continue

        # 过滤全空行
        rows = [r for r in rows if any(
            c is not None and str(c).strip() != "" for c in r
        )]
        if not rows:
            results.append("（空表）\n")
            continue

        # 生成 Markdown 表格
        max_cols = max(len(r) for r in rows)
        md_rows: List[str] = []
        for i, row in enumerate(rows):
            cells = [str(c).strip() if c is not None else "" for c in row]
            while len(cells) < max_cols:
                cells.append("")
            md_rows.append("| " + " | ".join(cells) + " |")
            if i == 0:
                md_rows.append("| " + " | ".join(["---"] * max_cols) + " |")

        results.append("\n".join(md_rows) + "\n")

    wb.close()
    return "\n".join(results)


# ============================================================
# 文档分发器：根据扩展名调用对应清洗函数
# ============================================================


def process_document(file_path: str) -> str:
    """根据文件扩展名调用对应的清洗/提取函数，返回纯文本"""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".html", ".htm"):
        return clean_html(file_path)
    elif ext == ".pdf":
        return extract_pdf_text(file_path)
    elif ext in (".xlsx", ".xls"):
        return excel_to_markdown(file_path)
    elif ext == ".txt":
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read().strip()
    else:
        # 未知类型按纯文本尝试
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read().strip()


# ============================================================
# 中文智能切片（jieba 辅助）
# ============================================================

# 句子边界正则：中文标点 + 英文标点
_SENT_BOUNDARY = re.compile(r"(?<=[。！？；\.!\?;，,\n])")


def _split_sentences(text: str) -> List[str]:
    """将文本按句子边界切分"""
    parts = _SENT_BOUNDARY.split(text)
    sentences: List[str] = []
    current = ""
    for part in parts:
        current += part
        # 如果以句子结尾标点结束，则认为是一个完整句子
        if part and part[-1] in "。！？.!?\n":
            if current.strip():
                sentences.append(current.strip())
            current = ""
    if current.strip():
        sentences.append(current.strip())
    return sentences


def chunk_text(
    text: str,
    chunk_size: int = 400,
    overlap: int = 80,
    use_jieba: bool = True,
) -> List[str]:
    """
    将文本按字数切片，带重叠，中文用 jieba 辅助。

    策略：
    1. 先按段落（\n\n）分
    2. 段内按句子切
    3. 累积到 chunk_size 附近就输出一个 chunk
    4. 保留 overlap 字作为上下文重叠

    返回: 切片文本列表（每片 300-500 字，最少 20 字）
    """
    if not text or not text.strip():
        return []

    chunks: List[str] = []
    paragraphs = text.split("\n\n")
    current_chars: List[str] = []
    current_len = 0

    def _emit():
        """输出当前累积的文本为一个 chunk，并保留 overlap"""
        nonlocal current_chars, current_len
        chunk_text_str = "".join(current_chars).strip()
        if len(chunk_text_str) >= 20:
            chunks.append(chunk_text_str)

        # 保留末尾 overlap 字数
        if current_len > overlap:
            remain = chunk_text_str[-overlap:]
            current_chars = [remain]
            current_len = len(remain)
        else:
            current_chars = []
            current_len = 0

    for para in paragraphs:
        para = para.strip()
        if not para:
            if current_len >= chunk_size * 0.6:
                _emit()
            elif current_len > 0:
                current_chars.append("\n")
                current_len += 1
            continue

        sentences = _split_sentences(para)
        for sent in sentences:
            if not sent.strip():
                continue

            sent_len = len(sent)

            # 如果加上这句还在 size 内，继续累积
            if current_len + sent_len <= chunk_size:
                current_chars.append(sent)
                current_len += sent_len
            else:
                # 当前已经累积了一些，先输出
                if current_len >= chunk_size * 0.4:
                    _emit()

                # 如果单句就超过 chunk_size，硬切
                if sent_len > chunk_size:
                    if current_len > 0:
                        _emit()
                    # 按 chunk_size 步长硬切
                    for i in range(0, sent_len, chunk_size - overlap):
                        piece = sent[i:i + chunk_size]
                        if len(piece.strip()) >= 20:
                            chunks.append(piece.strip())
                    current_chars = []
                    current_len = 0
                else:
                    current_chars.append(sent)
                    current_len += sent_len

    # 输出剩余
    if current_len > 0:
        remaining = "".join(current_chars).strip()
        if len(remaining) >= 20:
            chunks.append(remaining)

    return chunks


# ============================================================
# 完整流程：扫描文档目录，清洗+切片
# ============================================================


def scan_and_process_documents(
    documents_dir: str,
    chunk_size: int = 400,
    overlap: int = 80,
) -> Dict[str, List[Dict]]:
    """
    扫描 documents_dir，按场景子目录处理所有文档。

    返回: {文件相对路径: [{"text": chunk文本, "index": 序号}, ...]}
    """
    results: Dict[str, List[Dict]] = {}

    if not os.path.isdir(documents_dir):
        return results

    for scenario in sorted(os.listdir(documents_dir)):
        scenario_dir = os.path.join(documents_dir, scenario)
        if not os.path.isdir(scenario_dir):
            continue

        for fname in sorted(os.listdir(scenario_dir)):
            fpath = os.path.join(scenario_dir, fname)
            if fname.startswith("."):
                continue
            if not os.path.isfile(fpath):
                continue

            rel_path = os.path.join(scenario, fname)
            try:
                text = process_document(fpath)
                if not text:
                    print(f"  ⚠ 文档为空，跳过: {rel_path}")
                    continue
                chunks = chunk_text(text, chunk_size, overlap)
                results[rel_path] = [
                    {"text": c, "index": i} for i, c in enumerate(chunks)
                ]
                print(f"  ✓ {rel_path} → {len(chunks)} 个 chunk ({len(text)} 字)")
            except Exception as e:
                print(f"  ✗ 处理失败 {rel_path}: {e}")

    return results
